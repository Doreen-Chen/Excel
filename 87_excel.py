# 87. 處理 Excel 檔 (使用第三方套件流程示範)
# & 88. 處理 Excel 檔 (應用展示)
#   https://openpyxl.readthedocs.io/en/stable/


# Python 所有東西都是物件
### Key Classes  (from  https://openpyxl.readthedocs.io/en/stable/ )
### openpyxl.workbook.workbook.Workbook
### openpyxl.worksheet.worksheet.Worksheet
### openpyxl.cell.cell.Cell


# from 指定套件 ; import class
from openpyxl import Workbook
wb = Workbook()  # Workbook: 第1碼大寫為物件
# x = 5  # type : int ; 而 Workbook 的 type 為 Workbook


# grab the active worksheet
ws = wb.active


# Data can be assigned directly to cells
ws['A1'] = 42  # 42也是物件
ws['B1'] = 'Doreen'

# Rows can also be appended
ws.append([1, 2, 3])  # 清單[1, 2, 3]裝著3個物件

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")  # 使用 wb 的 save 功能